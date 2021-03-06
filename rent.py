import sys, os, time
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import user
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import time
import xlrd
from openpyxl import load_workbook 
from PIL import Image, ImageOps

username = user.username
passkey = user.password






# open an excel file
raw = xlrd.open_workbook('addresses1.xlsx')

# pulling the first sheet of the file
sheet = raw.sheet_by_index(0)

# calculating the number of addresses
count = len(sheet.col_values(0))-1

addresses = []

# formating and populating addresses array using data from address spreadsheet
for i in range(count):
    addresses.append( str(int(sheet.cell(i+1, 0).value)) + " " + str(sheet.cell(i+1, 1).value) + " " + str(sheet.cell(i+1, 2).value) + "," + " " + str(sheet.cell(i+1, 3).value) + "," + " " + str(sheet.cell(i+1, 4).value) + " " + str(int(sheet.cell(i+1, 5).value)) )



# selenium


PATH = "C:\Program Files (x86)\chromedriver.exe"


# options to open browser window fully maxmimized and page zoomed in
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
options.add_argument(f"--force-device-scale-factor={1.1}")


#defining driver for capturing resnt extimate chart
driver = webdriver.Chrome(PATH, options=options)


# opening website used to retrieve rent estimate chart
driver.get("https://www.gosection8.com/logreg.aspx?user=landlord")


# sending user info to log into website
email = driver.find_element_by_id("txtUsername")
email.send_keys(username)
password = driver.find_element_by_id("txtPassword")
password.send_keys(passkey)
password.send_keys(Keys.RETURN)


# going to rent estimate section of website
driver.get("https://www.gosection8.com/ll/ll_CompareRent.aspx")

agree = driver.find_element_by_id("MainContentPlaceHolder_MainContent2_disclosureButton")

agree.send_keys(Keys.RETURN)

# defining sizes the rent estimate chart is cropped into
size = (450, 360)
size1 = (85,55,185,300)
size2 = (185,55,255,300)
size3 = (255,55,330,300)
size4 = (330,55,430,300)

#array to store the file names of the cropped images 
imageList = []


# looping through the address list to enter them into the rent estimate website for different bedroom numbers and 
# capturing chart and cropping the chart into smaller images where each image contains only one bar from the chart
for strAddress in addresses:
    for bed in range(4):
        address = driver.find_element_by_id("MainContentPlaceHolder_MainContent2_AutocompleteAddress")
        beds = driver.find_element_by_id("MainContentPlaceHolder_MainContent2_BedroomCount")
        address.clear()
        beds.clear()
        address.send_keys(strAddress)
        beds.send_keys(bed+1)
        beds.send_keys(Keys.RETURN)  
        image = driver.find_element_by_id("MainContentPlaceHolder_MainContent2_compareRentChart")
        image.screenshot("./charts/"+strAddress.replace(" ", "").replace(",", "")+"-"+str(bed+1)+".png")
        openImg = Image.open("./charts/"+strAddress.replace(" ", "").replace(",", "")+"-"+str(bed+1)+".png")
        resizeImg = ImageOps.fit(openImg, size, Image.ANTIALIAS)
        cropImg1 = resizeImg.crop(size1)
        cropImg2 = resizeImg.crop(size2)
        cropImg3 = resizeImg.crop(size3)
        cropImg4 = resizeImg.crop(size4)

        cropImg1.save("./charts/"+strAddress.replace(" ", "").replace(",", "")+"-"+str(bed+1)+"-1"+".png")
        imageList.append("/charts/"+strAddress.replace(" ", "").replace(",", "")+"-"+str(bed+1)+"-1"+".png")

        cropImg2.save("./charts/"+strAddress.replace(" ", "").replace(",", "")+"-"+str(bed+1)+"-2"+".png")
        imageList.append("/charts/"+strAddress.replace(" ", "").replace(",", "")+"-"+str(bed+1)+"-2"+".png")

        cropImg3.save("./charts/"+strAddress.replace(" ", "").replace(",", "")+"-"+str(bed+1)+"-3"+".png")
        imageList.append("/charts/"+strAddress.replace(" ", "").replace(",", "")+"-"+str(bed+1)+"-3"+".png")

        cropImg4.save("./charts/"+strAddress.replace(" ", "").replace(",", "")+"-"+str(bed+1)+"-4"+".png")
        imageList.append("/charts/"+strAddress.replace(" ", "").replace(",", "")+"-"+str(bed+1)+"-4"+".png")

# closing rent estimate website chrome window
driver.close()
driver.quit()

# opening the spreadsheet, storing the filenames of the cropped images and saving the spreadsheet
wb = load_workbook(filename="addresses1.xlsx")
ws = wb.worksheets[0]

for index, fileName in enumerate(imageList):
    ws.cell(row=(index//16)+2, column=23+(index%16)).value = fileName

wb.save(filename="addresses1.xlsx")



#retrieving the cropped image file names from spreadsheet and storing them into an array
# to use the names to call the images and feed them into Optical Character Recognition website
fileList = []

wb = load_workbook(filename="addresses1.xlsx")
ws = wb.worksheets[0]

# get number of streest addresses from spreadsheet
numStreet = len(ws['A'])-1

for row in range(numStreet):
    for col in range(16):
        fileList.append( ws.cell(row=row+2, column=23+col).value )


# defining an array of drivers to open multiple chrome windows at once to save time
driverOCR = []

# populating drivers array
for i in range(16):
    driverOCR.append(webdriver.Chrome(PATH, options=options))

# opening optical character recognition (OCR) website on each chrome window 
for driverSingle in driverOCR:
    driverSingle.get("https://brandfolder.com/workbench/extract-text-from-image")


# looping through the cropped images to feed them into the OCR website and retrieve rent estimate data
# and feed the data into the spreadsheet
for j in range(len(fileList)//16):
    time.sleep(5)
    upload = []
    rentEst = []
    for k in range(16):
        print("k=", k)
        try:
            #capturing the input element of the OCR website to feed image
            upload.append(driverOCR[k].find_element_by_class_name("fsp-drop-pane__input"))
        except:
            # in case capturing input element fails this esception reloads website and tries capturing input element again
            print("exception error when appending upload and k=", k)
            driverOCR[k].get("https://brandfolder.com/workbench/extract-text-from-image")
            time.sleep(2)
            upload.append(driverOCR[k].find_element_by_class_name("fsp-drop-pane__input"))

        # clear the upload input element and send the cropped image
        upload[k].clear()
        upload[k].send_keys(os.getcwd() + fileList[16*j + k])
    
    # wait afew moments until the OCR website processes images and sipts out text data
    time.sleep(5)

    # loop through the open chrome windows to capture the output text data
    for l in range(16):

        # capture the text element of the OCR website
        text = driverOCR[l].find_element_by_id("extracted_text").text

        # in case the image is not uploaded and there is no output text try reloading the page 
        # and upload image again and capture text data
        while(len(text) == 0):
            print("when text is null l=", l)
            driverOCR[l].get("https://brandfolder.com/workbench/extract-text-from-image")
            time.sleep(2)
            uploadS = driverOCR[l].find_element_by_class_name("fsp-drop-pane__input")
            uploadS.clear()
            uploadS.send_keys(os.getcwd() + fileList[16*j + l])
            time.sleep(5)
            text = driverOCR[l].find_element_by_id("extracted_text").text
        
        # add text data to rent estimate array
        rentEst.append(text)

        # refresh the page to prepare for feeding the cropped images for the next street address 
        driverOCR[l].refresh()

    # open spreadsheet and save the rent estimate data array into the spreadsheet
    wb = load_workbook(filename="addresses1.xlsx")
    ws = wb.worksheets[0]

    for index, estimate in enumerate(rentEst):
        ws.cell(row=j+2, column=index+7).value = estimate

    wb.save(filename="addresses1.xlsx")


# once all rent estimate data is saved into the spreadhseet close and quit all drivers and chrome windows
for driverS in driverOCR:
    driverS.close()
    driverS.quit()

